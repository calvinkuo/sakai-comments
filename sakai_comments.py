# Calvin Kuo (clk3sx)
# 2021-03-28

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as col

PATH_FORMAT = '{}/{}({})/comments.txt'
TSV_SEPARATOR = '\t'


class Comment:
    def __init__(self, computing_id: str, last_name: str, first_name: str, comment: str):
        self.computing_id = computing_id
        self.last_name = last_name
        self.first_name = first_name
        self.name = '{}, {}'.format(last_name, first_name) \
            if last_name is not None and first_name is not None else '{},'.format(computing_id)
        self.comment = comment.replace('\\n', '\n') if comment is not None else ''

    def write(self, root_path):
        with open(PATH_FORMAT.format(root_path, self.name, self.computing_id),
                  'w',
                  encoding='utf-8') as f:
            f.write(self.comment)

    def __repr__(self):
        return "Comment({}, {}, {}, {})".format(repr(self.computing_id),
                                                repr(self.last_name),
                                                repr(self.first_name),
                                                repr(self.comment))


def parse_xls(path, comment_key):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    print("Worksheet loaded")

    comments_list = parse_ws(ws, comment_key)

    wb.close()
    print("All comments loaded")
    return comments_list


def parse_ws(ws, comment_key):
    # get indices of column headers
    column_index = {}
    for c in range(1, 20):
        key = ws["{}1".format(col(c))].value
        if key is None:
            break
        column_index[key.strip()] = col(c)

    # load each submission
    comments_list = []
    for r in range(2, 400):
        computing_id = ws["{}{}".format(column_index['ID'], r)].value
        last_name = ws["{}{}".format(column_index['Last Name'], r)].value
        first_name = ws["{}{}".format(column_index['First Name'], r)].value
        comment = ws["{}{}".format(column_index[comment_key], r)].value
        if computing_id is None:
            break
        comments_list.append(Comment(computing_id, last_name, first_name, comment))

    return comments_list


def parse_tsv(path, comment_key):
    comments_list = []
    with open(path, 'r') as f:
        # get indices of column headers
        header = f.readline().rstrip(' \n').split(TSV_SEPARATOR)
        index_id = header.index('ID')
        index_last_name = header.index('Last Name')
        index_first_name = header.index('First Name')
        index_feedback = header.index(comment_key)

        # load each submission
        for line in f.readlines():
            values = line.rstrip(' \n').split(TSV_SEPARATOR)
            comments_list.append(Comment(values[index_id],
                                         values[index_last_name],
                                         values[index_first_name],
                                         values[index_feedback]))

    print("All comments loaded")
    return comments_list


def write_comments(comments_list, path):
    for comment in comments_list:
        comment.write(path)
    print("All comments written")


def run(comments_path, folder_path, header_comment):
    # strip any quotes from xls_path
    if comments_path.startswith('"') and comments_path.endswith('"'):
        comments_path = comments_path[1:-1]
    if folder_path.startswith('"') and folder_path.endswith('"'):
        folder_path = folder_path[1:-1]

    write_comments(parse_xls(comments_path, header_comment), folder_path)


if __name__ == "__main__":
    comments = input("File xls_path for grades.xls? ").strip()
    folder = input("Assignment xls_path (the folder with each student's folder)? ").strip()
    header = input("Column header for comments? ").strip()
    run(comments, folder, header)
